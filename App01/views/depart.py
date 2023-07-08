from django.shortcuts import render, HttpResponse, redirect
from App01.models import UserInfo, Department, PrettyNumber
from App01.utils.pagination import Pagination
# Create your views here.

""" 部门列表 """


def depart_list(request):

    list = Department.objects.all()
    obj = Pagination(request, list)

    content = {
        'list': obj.query_set,
        'page_string': obj.html(),
    }

    return render(request, 'depart_list.html', content)


''' 新建部门 '''


def depart_add(request):
    if request.method == 'GET':
        return render(request, 'depart_add.html')
    else:
        title = request.POST.get('title')
        Department.objects.create(title=title)

        return redirect('/depart/list/')


def depart_multi(request):
    file_object = request.FILES.get('file')
    print(type(file_object))
    from openpyxl import load_workbook
    # 读取文件
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        title = row[0].value
        if not Department.objects.filter(title=title).exists():
            Department.objects.create(title=title)

    return redirect('/depart/list/')


""" 删除部门 """


def depart_delete(request):
    nid = request.GET.get('nid')

    Department.objects.filter(id=nid).delete()

    return redirect('/depart/list/')


""" 编辑部门 """


def depart_edit(request, nid):
    if request.method == 'GET':
        title = Department.objects.filter(id=nid).first().title
        return render(request, 'depart_edit.html', {'title': title})
    else:
        title = request.POST.get('title')
        Department.objects.filter(id=nid).update(title=title)
        return redirect('/depart/list/')

